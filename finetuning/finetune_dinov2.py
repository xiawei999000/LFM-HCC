import argparse
import os
import time
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.utils.data as data
import pandas as pd
import numpy as np
import SimpleITK as sitk
from torch.optim import lr_scheduler
from sklearn.metrics import roc_auc_score
from tqdm import tqdm
from timm.models.vision_transformer import vit_small_patch8_224
import openpyxl


# ============================================================
# Configuration
# ============================================================
class Config:
    def __init__(self, args):
        # Data paths
        self.COHORT_A_EXCEL = args.cohort_a_excel
        self.COHORT_A_AP_DIR = args.cohort_a_ap_dir
        self.COHORT_A_PV_DIR = args.cohort_a_pv_dir
        self.COHORT_B_EXCEL = args.cohort_b_excel
        self.COHORT_B_AP_DIR = args.cohort_b_ap_dir
        self.COHORT_B_PV_DIR = args.cohort_b_pv_dir

        # Pretrained weights path
        self.PRETRAINED_WEIGHTS = args.pretrained_weights
        pretrain_tag = os.path.splitext(os.path.basename(args.pretrained_weights))[0] if args.pretrained_weights else "imagenet"

        # Output
        self.OUTPUT_DIR = args.output_dir
        self.TRAIN_LOG_DIR = os.path.join(self.OUTPUT_DIR, "train_log")

        # Training hyperparams
        self.BATCH_SIZE = args.batch_size
        self.LR_LIST = [float(x) for x in args.lr_list.split(",")]
        self.LR = self.LR_LIST[0]
        self.BACKBONE_LR = self.LR
        self.HEAD_LR = self.LR
        self.WEIGHT_DECAY = args.weight_decay
        self.EPOCHS = args.epochs
        self.PATIENCE = args.patience

        self.EMBED_DIM = 384
        self.LABEL_NAMES = ['MVI', 'PathologicalGrade']


def set_lr(cfg, lr):
    cfg.LR = lr
    cfg.BACKBONE_LR = lr
    cfg.HEAD_LR = lr


def lr_to_tag(lr):
    lr_str = f"{lr:.0e}" if lr < 1 else f"{lr:g}"
    lr_str = lr_str.replace("+", "")
    return f"lr_{lr_str}"


def prepare_run_dirs(cfg, lr):
    lr_tag = lr_to_tag(lr)
    weight_dir = os.path.join(cfg.OUTPUT_DIR, lr_tag)
    os.makedirs(weight_dir, exist_ok=True)
    os.makedirs(cfg.TRAIN_LOG_DIR, exist_ok=True)
    return lr_tag, weight_dir, cfg.TRAIN_LOG_DIR


def update_lr_summary(train_log_dir, summary_row):
    csv_path = os.path.join(train_log_dir, "lr_summary.csv")
    xlsx_path = os.path.join(train_log_dir, "lr_summary.xlsx")
    new_df = pd.DataFrame([summary_row])

    if os.path.exists(csv_path):
        old_df = pd.read_csv(csv_path)
        if 'lr_tag' in old_df.columns:
            old_df = old_df[old_df['lr_tag'] != summary_row['lr_tag']]
        summary_df = pd.concat([old_df, new_df], ignore_index=True)
    else:
        summary_df = new_df

    summary_df = summary_df.sort_values(by="lr", ascending=True).reset_index(drop=True)
    summary_df.to_csv(csv_path, index=False)
    summary_df.to_excel(xlsx_path, index=False)


def select_best_lr_and_save_report(train_log_dir):
    csv_path = os.path.join(train_log_dir, "lr_summary.csv")
    if not os.path.exists(csv_path):
        print("[!] lr_summary.csv not found, cannot auto-select best LR.")
        return None

    df = pd.read_csv(csv_path)
    if len(df) == 0:
        print("[!] lr_summary.csv is empty, cannot auto-select best LR.")
        return None

    # Composite metric: average of val MVI and Grade AUC; tie-break by val loss
    df["best_val_auc_mean"] = (df["best_val_mvi"] + df["best_val_grade"]) / 2.0

    best_overall = df.sort_values(
        by=["best_val_auc_mean", "best_val_loss", "best_val_mvi", "best_val_grade"],
        ascending=[False, True, False, False]
    ).iloc[0]

    best_mvi = df.sort_values(
        by=["best_val_mvi", "best_val_loss"],
        ascending=[False, True]
    ).iloc[0]

    best_grade = df.sort_values(
        by=["best_val_grade", "best_val_loss"],
        ascending=[False, True]
    ).iloc[0]

    report_rows = [
        {
            "category": "overall_best_lr",
            "criterion": "max((best_val_mvi + best_val_grade)/2), tie -> min(best_val_loss)",
            **{k: best_overall[k] for k in best_overall.index}
        },
        {
            "category": "best_mvi_lr",
            "criterion": "max(best_val_mvi), tie -> min(best_val_loss)",
            **{k: best_mvi[k] for k in best_mvi.index}
        },
        {
            "category": "best_grade_lr",
            "criterion": "max(best_val_grade), tie -> min(best_val_loss)",
            **{k: best_grade[k] for k in best_grade.index}
        }
    ]

    report_df = pd.DataFrame(report_rows)
    report_csv = os.path.join(train_log_dir, "best_lr_report.csv")
    report_xlsx = os.path.join(train_log_dir, "best_lr_report.xlsx")
    report_txt = os.path.join(train_log_dir, "best_lr_report.txt")

    report_df.to_csv(report_csv, index=False)
    report_df.to_excel(report_xlsx, index=False)

    with open(report_txt, "w", encoding="utf-8") as f:
        f.write("Auto-Selected Best LR Results\n")
        f.write("=" * 60 + "\n\n")
        for cat in ["overall_best_lr", "best_mvi_lr", "best_grade_lr"]:
            row = report_df[report_df["category"] == cat].iloc[0]
            f.write(f"[{cat}]\n")
            f.write(f"  LR tag: {row['lr_tag']}, LR: {row['lr']}\n")
            f.write(f"  Val MVI: {row['best_val_mvi']:.4f}, Val Grade: {row['best_val_grade']:.4f}\n\n")

    print("\n" + "=" * 70)
    print("Auto-Selected Best LR")
    print("=" * 70)
    best = best_overall
    print(f"[Overall] LR={best['lr']}, tag={best['lr_tag']}")
    print(f"  Val Mean AUC: {best['best_val_auc_mean']:.4f}, Loss: {best['best_val_loss']:.6f}")
    print(f"  Val MVI: {best['best_val_mvi']:.4f}, Grade: {best['best_val_grade']:.4f}")
    print(f"  Weights: {best['weight_dir']}")
    print("-" * 70)
    print(f"[MVI] LR={best_mvi['lr']}, Best Val AUC: {best_mvi['best_val_mvi']:.4f}")
    print(f"[Grade] LR={best_grade['lr']}, Best Val AUC: {best_grade['best_val_grade']:.4f}")
    print("-" * 70)
    print(f"Reports saved to: {report_csv}, {report_xlsx}")
    print("=" * 70)

    return {
        "overall_best_lr": best_overall.to_dict(),
        "best_mvi_lr": best_mvi.to_dict(),
        "best_grade_lr": best_grade.to_dict(),
        "report_csv": report_csv,
        "report_xlsx": report_xlsx,
        "report_txt": report_txt,
    }


# ============================================================
# Dataset — in-memory preloading
# ============================================================
class HCCDataset(data.Dataset):
    def __init__(self, df, ap_dir, pv_dir, desc="Loading Data",
                 fallback_ap_dir=None, fallback_pv_dir=None):
        self.df = df.reset_index(drop=True)
        self.data_cache = []

        print(f"[*] Loading {desc} into RAM...")
        for idx in tqdm(range(len(self.df)), desc=desc):
            row = self.df.iloc[idx]
            pid = str(row['ID'])
            label = torch.tensor(
                [float(row['MVI']), float(row['PathologicalGrade'])],
                dtype=torch.float32
            )

            ap_path = os.path.join(ap_dir, f"{pid}.nii.gz")
            pv_path = os.path.join(pv_dir, f"{pid}.nii.gz")

            if not os.path.exists(ap_path) and fallback_ap_dir:
                ap_path = os.path.join(fallback_ap_dir, f"{pid}.nii.gz")
            if not os.path.exists(pv_path) and fallback_pv_dir:
                pv_path = os.path.join(fallback_pv_dir, f"{pid}.nii.gz")

            self.data_cache.append({
                'ap': self._load_nii(ap_path),
                'pv': self._load_nii(pv_path),
                'label': label,
                'pid': pid
            })

    def __len__(self):
        return len(self.data_cache)

    @staticmethod
    def _load_nii(file_path):
        img = sitk.ReadImage(file_path)
        arr = sitk.GetArrayFromImage(img).astype(np.float32) / 255.0
        return torch.from_numpy(arr).unsqueeze(1)

    def __getitem__(self, idx):
        item = self.data_cache[idx]
        return item['ap'], item['pv'], item['label'], item['pid']


# ============================================================
# Collate function — keep variable-length lists
# ============================================================
def my_collate(batch):
    ap_list, pv_list, labels, pids = [], [], [], []
    for ap, pv, label, pid in batch:
        ap_list.append(ap)
        pv_list.append(pv)
        labels.append(label)
        pids.append(pid)
    labels = torch.stack(labels, dim=0)
    return ap_list, pv_list, labels, pids


# ============================================================
# Data splitting
# ============================================================
def prepare_data_splits(cfg):
    def filter_valid(df, ap_dir, pv_dir):
        mask = [
            os.path.exists(os.path.join(ap_dir, f"{pid}.nii.gz")) and
            os.path.exists(os.path.join(pv_dir, f"{pid}.nii.gz"))
            for pid in df['ID'].astype(str)
        ]
        return df[mask].reset_index(drop=True)

    df_cohort_a = pd.read_excel(cfg.COHORT_A_EXCEL)
    df_cohort_a = filter_valid(df_cohort_a, cfg.COHORT_A_AP_DIR, cfg.COHORT_A_PV_DIR)
    df_cohort_a['set_split'] = df_cohort_a['set_split'].astype(str).str.strip().str.lower()

    train_df = df_cohort_a[df_cohort_a['set_split'] == 'train']
    val_df = df_cohort_a[df_cohort_a['set_split'] == 'val']
    ext1_df = df_cohort_a[df_cohort_a['set_split'] == 'ext1']

    df_cohort_b = pd.read_excel(cfg.COHORT_B_EXCEL)
    df_cohort_b = filter_valid(df_cohort_b, cfg.COHORT_B_AP_DIR, cfg.COHORT_B_PV_DIR)

    print(f"[*] Split: Train={len(train_df)}, Val={len(val_df)}, "
          f"Cohort A ext1={len(ext1_df)}, Cohort B ext2={len(df_cohort_b)}")

    train_ds = HCCDataset(train_df, cfg.COHORT_A_AP_DIR, cfg.COHORT_A_PV_DIR, desc="Train Set",
                          fallback_ap_dir=cfg.COHORT_B_AP_DIR, fallback_pv_dir=cfg.COHORT_B_PV_DIR)
    val_ds = HCCDataset(val_df, cfg.COHORT_A_AP_DIR, cfg.COHORT_A_PV_DIR, desc="Val Set")
    ext1_ds = HCCDataset(ext1_df, cfg.COHORT_A_AP_DIR, cfg.COHORT_A_PV_DIR, desc="Cohort A (ext1)")
    ext2_ds = HCCDataset(df_cohort_b, cfg.COHORT_B_AP_DIR, cfg.COHORT_B_PV_DIR, desc="Cohort B (ext2)")

    return train_ds, val_ds, ext1_ds, ext2_ds


# ============================================================
# Model components
# ============================================================
class CrossPhaseAttention(nn.Module):
    def __init__(self, dim, num_heads=4, dropout=0.1):
        super().__init__()
        self.norm_ap = nn.LayerNorm(dim)
        self.norm_pv = nn.LayerNorm(dim)
        self.ap_query_pv = nn.MultiheadAttention(
            embed_dim=dim, num_heads=num_heads, dropout=dropout, batch_first=True)
        self.pv_query_ap = nn.MultiheadAttention(
            embed_dim=dim, num_heads=num_heads, dropout=dropout, batch_first=True)

    def forward(self, ap_feat, pv_feat):
        ap_norm = self.norm_ap(ap_feat)
        pv_norm = self.norm_pv(pv_feat)
        ap_cross, _ = self.ap_query_pv(query=ap_norm, key=pv_norm, value=pv_norm)
        pv_cross, _ = self.pv_query_ap(query=pv_norm, key=ap_norm, value=ap_norm)
        return ap_feat + ap_cross, pv_feat + pv_cross


class GatedAttention(nn.Module):
    def __init__(self, in_dim, hidden_dim=256):
        super().__init__()
        self.V = nn.Sequential(nn.Linear(in_dim, hidden_dim), nn.Tanh())
        self.U = nn.Sequential(nn.Linear(in_dim, hidden_dim), nn.Sigmoid())
        self.w = nn.Linear(hidden_dim, 1)

    def forward(self, x):
        a = self.w(self.V(x) * self.U(x))
        a = torch.softmax(a, dim=0)
        z = torch.sum(x * a, dim=0, keepdim=True)
        return z, a


class MultiTaskMILModel(nn.Module):
    def __init__(self, cfg):
        super().__init__()
        self.backbone = vit_small_patch8_224(
            img_size=128, in_chans=1, num_classes=0,
            pretrained=False, dynamic_img_size=True, init_values=1e-5,
        )
        self._load_pretrained(cfg.PRETRAINED_WEIGHTS)

        D = cfg.EMBED_DIM
        fusion_dim = D * 2

        self.cross_phase_attn = CrossPhaseAttention(dim=D, num_heads=4, dropout=0.1)
        self.attention = GatedAttention(fusion_dim)

        self.head_mvi = nn.Sequential(
            nn.Linear(fusion_dim, fusion_dim // 2), nn.ReLU(),
            nn.Dropout(0.3), nn.Linear(fusion_dim // 2, 1))
        self.head_grade = nn.Sequential(
            nn.Linear(fusion_dim, fusion_dim // 2), nn.ReLU(),
            nn.Dropout(0.3), nn.Linear(fusion_dim // 2, 1))

        self.log_vars = nn.Parameter(torch.zeros(2))

    def _load_pretrained(self, weight_path):
        if not os.path.exists(weight_path):
            print(f"[!] Pretrained weights not found: {weight_path}, using random init.")
            return
        ckpt = torch.load(weight_path, map_location='cpu', weights_only=False)
        full = ckpt.get('model_state_dict', ckpt)
        sub = {k.replace('teacher_backbone.vit.', ''): v
               for k, v in full.items() if k.startswith('teacher_backbone.vit.')}
        msg = self.backbone.load_state_dict(sub, strict=False)
        print(f"[*] Loaded pretrained weights. "
              f"missing={len(msg.missing_keys)}, unexpected={len(msg.unexpected_keys)}")

    def forward_one_patient(self, ap_tensor, pv_tensor):
        feat_ap = self.backbone(ap_tensor)
        feat_pv = self.backbone(pv_tensor)
        feat_ap = feat_ap.unsqueeze(0)
        feat_pv = feat_pv.unsqueeze(0)
        feat_ap, feat_pv = self.cross_phase_attn(feat_ap, feat_pv)
        feat_ap = feat_ap.squeeze(0)
        feat_pv = feat_pv.squeeze(0)
        feat_fused = torch.cat([feat_ap, feat_pv], dim=1)
        z, _ = self.attention(feat_fused)
        out_mvi = self.head_mvi(z).view(-1)
        out_grade = self.head_grade(z).view(-1)
        return torch.cat([out_mvi, out_grade], dim=0)

    def forward(self, ap_list, pv_list):
        outputs = [self.forward_one_patient(ap, pv) for ap, pv in zip(ap_list, pv_list)]
        return torch.stack(outputs, dim=0)


# ============================================================
# Logging
# ============================================================
def log_create():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('results', 0)
    headers = [
        'lr', 'epoch', 'backbone_lr', 'head_lr',
        'train_loss', 'val_loss', 'weight_mvi', 'weight_grade',
        'Train_MVI_AUC', 'Train_Grade_AUC',
        'Val_MVI_AUC', 'Val_Grade_AUC',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    return wb, ws


# ============================================================
# Loss function
# ============================================================
def compute_multitask_uncertainty_loss(logits, labels, log_vars):
    pred_mvi = logits[:, 0].view(-1)
    pred_grade = logits[:, 1].view(-1)
    target_mvi = labels[:, 0].view(-1)
    target_grade = labels[:, 1].view(-1)

    loss_mvi = F.binary_cross_entropy_with_logits(pred_mvi, target_mvi)
    loss_grade = F.binary_cross_entropy_with_logits(pred_grade, target_grade)

    precision_mvi = torch.exp(-log_vars[0])
    precision_grade = torch.exp(-log_vars[1])

    total_loss = (precision_mvi * loss_mvi + log_vars[0]
                  + precision_grade * loss_grade + log_vars[1])
    return total_loss, loss_mvi.detach(), loss_grade.detach()


# ============================================================
# Evaluation
# ============================================================
def evaluate(model, dataloader, device):
    model.eval()
    all_probs, all_labels = [], []
    total_loss = 0.0
    n_batches = 0

    with torch.no_grad():
        for ap_list, pv_list, labels, _ in dataloader:
            ap_list = [t.to(device) for t in ap_list]
            pv_list = [t.to(device) for t in pv_list]
            labels = labels.to(device)

            logits = model(ap_list, pv_list)
            loss, _, _ = compute_multitask_uncertainty_loss(logits, labels, model.log_vars)

            all_probs.append(torch.sigmoid(logits).cpu().numpy())
            all_labels.append(labels.cpu().numpy())
            total_loss += loss.item()
            n_batches += 1

    avg_loss = total_loss / max(n_batches, 1)
    all_probs = np.concatenate(all_probs, axis=0)
    all_labels = np.concatenate(all_labels, axis=0)

    auc_scores = []
    for i in range(all_labels.shape[1]):
        try:
            auc = roc_auc_score(all_labels[:, i], all_probs[:, i])
        except ValueError:
            auc = 0.5
        auc_scores.append(round(auc, 4))

    return auc_scores, avg_loss


# ============================================================
# Training loop
# ============================================================
def train_model(cfg, model, dataloaders, device, weight_dir, train_log_dir, lr_tag):
    optimizer = torch.optim.AdamW([
        {'params': model.backbone.parameters(), 'lr': cfg.BACKBONE_LR},
        {'params': model.cross_phase_attn.parameters(), 'lr': cfg.HEAD_LR},
        {'params': model.attention.parameters(), 'lr': cfg.HEAD_LR},
        {'params': model.head_mvi.parameters(), 'lr': cfg.HEAD_LR},
        {'params': model.head_grade.parameters(), 'lr': cfg.HEAD_LR},
        {'params': [model.log_vars], 'lr': cfg.HEAD_LR},
    ], weight_decay=cfg.WEIGHT_DECAY)

    scheduler = lr_scheduler.CosineAnnealingLR(optimizer, T_max=cfg.EPOCHS, eta_min=1e-6)

    log_wb, log_ws = log_create()
    log_path = os.path.join(train_log_dir, f'training_log_{lr_tag}.xlsx')
    csv_path = os.path.join(train_log_dir, f'training_metrics_{lr_tag}.csv')
    csv_rows = []

    best_val_mvi = 0.0
    best_val_grade = 0.0
    best_val_loss = float('inf')
    patience_counter = 0
    last_epoch = 0

    best_mvi_record = {'epoch': 0, 'val': 0.0, 'filename': ''}
    best_grade_record = {'epoch': 0, 'val': 0.0, 'filename': ''}

    print("=" * 70)
    print(f"Training | LR={cfg.LR:.2e} | backbone unfrozen | Cross-Attn | Uncertainty Loss")
    print("=" * 70)

    for epoch in range(1, cfg.EPOCHS + 1):
        last_epoch = epoch
        epoch_start = time.time()
        model.train()

        running_loss = 0.0
        n_batches = 0

        for ap_list, pv_list, labels, _ in tqdm(
            dataloaders['train'], desc=f"LR {cfg.LR:.2e} | Epoch {epoch:03d} Train"):
            ap_list = [t.to(device) for t in ap_list]
            pv_list = [t.to(device) for t in pv_list]
            labels = labels.to(device)

            optimizer.zero_grad()
            logits = model(ap_list, pv_list)
            loss, _, _ = compute_multitask_uncertainty_loss(logits, labels, model.log_vars)
            loss.backward()
            optimizer.step()

            running_loss += loss.item()
            n_batches += 1

        avg_train_loss = running_loss / max(n_batches, 1)

        auc_train, _ = evaluate(model, dataloaders['train'], device)
        auc_val, val_loss = evaluate(model, dataloaders['val'], device)

        scheduler.step()

        w_mvi = torch.exp(-model.log_vars[0]).item()
        w_grade = torch.exp(-model.log_vars[1]).item()
        backbone_lr = optimizer.param_groups[0]['lr']
        head_lr = optimizer.param_groups[1]['lr']

        print(f"\n[LR {cfg.LR:.2e} | Epoch {epoch:03d}] "
              f"TrainLoss={avg_train_loss:.4f}  ValLoss={val_loss:.4f}  "
              f"W_MVI={w_mvi:.3f}  W_Grade={w_grade:.3f}  "
              f"BB_LR={backbone_lr:.2e}  HEAD_LR={head_lr:.2e}  "
              f"Time={time.time() - epoch_start:.1f}s")
        print(f"  Train  MVI={auc_train[0]:.4f}  Grade={auc_train[1]:.4f}")
        print(f"  Val    MVI={auc_val[0]:.4f}  Grade={auc_val[1]:.4f}")

        row_data = [
            cfg.LR, epoch, backbone_lr, head_lr,
            avg_train_loss, val_loss, w_mvi, w_grade,
            auc_train[0], auc_train[1],
            auc_val[0], auc_val[1],
        ]
        log_ws.append(row_data)
        log_wb.save(log_path)

        csv_rows.append({
            'LR': cfg.LR, 'Epoch': epoch,
            'Train_Loss': avg_train_loss, 'Val_Loss': val_loss,
            'Weight_MVI': w_mvi, 'Weight_Grade': w_grade,
            'Train_MVI': auc_train[0], 'Train_Grade': auc_train[1],
            'Val_MVI': auc_val[0], 'Val_Grade': auc_val[1],
        })
        pd.DataFrame(csv_rows).to_csv(csv_path, index=False)

        saved_any = False

        current_mvi_rnd = round(auc_val[0], 3)
        if current_mvi_rnd > best_val_mvi:
            best_val_mvi = current_mvi_rnd
            save_name = f"best_MVI_val{current_mvi_rnd:.3f}_epoch{epoch}.pth"
            torch.save(model.state_dict(), os.path.join(weight_dir, save_name))
            print(f"  >>> Saved best MVI model: {save_name}")
            best_mvi_record = {
                'epoch': epoch, 'val': auc_val[0], 'filename': save_name}
            saved_any = True

        current_grade_rnd = round(auc_val[1], 3)
        if current_grade_rnd > best_val_grade:
            best_val_grade = current_grade_rnd
            save_name = f"best_Grade_val{current_grade_rnd:.3f}_epoch{epoch}.pth"
            torch.save(model.state_dict(), os.path.join(weight_dir, save_name))
            print(f"  >>> Saved best Grade model: {save_name}")
            best_grade_record = {
                'epoch': epoch, 'val': auc_val[1], 'filename': save_name}
            saved_any = True

        if not saved_any:
            print("  --- Val AUC (MVI/Grade) did not improve, skipping save ---")

        if val_loss < best_val_loss:
            best_val_loss = val_loss
            patience_counter = 0
        else:
            patience_counter += 1
            print(f"  --- EarlyStopping counter: {patience_counter}/{cfg.PATIENCE} ---")

        if patience_counter >= cfg.PATIENCE:
            print(f"[!] Early stopping triggered at epoch {epoch}")
            break

    summary_row = {
        'lr_tag': lr_tag, 'lr': cfg.LR, 'epochs_ran': last_epoch,
        'best_val_loss': best_val_loss,
        'best_mvi_epoch': best_mvi_record['epoch'],
        'best_val_mvi': best_mvi_record['val'],
        'best_mvi_weight': best_mvi_record['filename'],
        'best_grade_epoch': best_grade_record['epoch'],
        'best_val_grade': best_grade_record['val'],
        'best_grade_weight': best_grade_record['filename'],
        'best_val_auc_mean': (best_mvi_record['val'] + best_grade_record['val']) / 2.0,
        'weight_dir': weight_dir,
        'epoch_log_xlsx': log_path, 'epoch_log_csv': csv_path,
    }

    print(f"\nTraining finished. LR={cfg.LR:.2e}")
    print(f"  Weights: {weight_dir}")
    print(f"  Logs: {train_log_dir}")
    return summary_row


# ============================================================
# Inference — predict per-patient scores
# ============================================================
@torch.no_grad()
def predict_cohort(model, dataloader, device):
    model.eval()
    all_probs, all_labels, all_pids = [], [], []

    for ap_list, pv_list, labels, pids in dataloader:
        ap_list = [t.to(device) for t in ap_list]
        pv_list = [t.to(device) for t in pv_list]

        logits = model(ap_list, pv_list)
        probs = torch.sigmoid(logits).cpu().numpy()

        all_probs.append(probs)
        all_labels.append(labels.cpu().numpy())
        all_pids.extend(pids)

    probs = np.concatenate(all_probs, axis=0)
    labels = np.concatenate(all_labels, axis=0)

    df = pd.DataFrame({
        "ID": [int(p) for p in all_pids],
        "MVI_label": labels[:, 0].astype(int),
        "MVI_prediction": (probs[:, 0] >= 0.5).astype(int),
        "MVI_prob": probs[:, 0].round(4),
        "PathologicalGrade_label": labels[:, 1].astype(int),
        "PathologicalGrade_prediction": (probs[:, 1] >= 0.5).astype(int),
        "PathologicalGrade_prob": probs[:, 1].round(4),
    })
    return df


def run_inference(cfg, best_result, device):
    weight_dir = best_result["overall_best_lr"]["weight_dir"]
    mvi_weight = best_result["overall_best_lr"]["best_mvi_weight"]
    checkpoint = os.path.join(weight_dir, mvi_weight)
    print(f"\n[*] Loading best model: {checkpoint}")

    model = MultiTaskMILModel(cfg).to(device)
    model.load_state_dict(torch.load(checkpoint, map_location=device))

    for cohort, ap_dir, pv_dir, excel_path, out_name in [
        ("Cohort A", cfg.COHORT_A_AP_DIR, cfg.COHORT_A_PV_DIR, cfg.COHORT_A_EXCEL, "predictions_cohort_a.xlsx"),
        ("Cohort B", cfg.COHORT_B_AP_DIR, cfg.COHORT_B_PV_DIR, cfg.COHORT_B_EXCEL, "predictions_cohort_b.xlsx"),
    ]:
        df_meta = pd.read_excel(excel_path)
        filenames = set(n.replace(".nii.gz", "") for n in os.listdir(ap_dir))
        valid = df_meta["ID"].astype(str).isin(filenames)
        df_meta = df_meta[valid].reset_index(drop=True)
        print(f"[*] {cohort}: {len(df_meta)} patients with images")

        ds = HCCDataset(df_meta, ap_dir, pv_dir, desc=f"{cohort} Inference")
        loader = data.DataLoader(
            ds, batch_size=cfg.BATCH_SIZE, shuffle=False,
            collate_fn=my_collate, num_workers=0, pin_memory=True)

        pred_df = predict_cohort(model, loader, device)
        out_path = os.path.join(cfg.OUTPUT_DIR, out_name)
        pred_df.to_excel(out_path, index=False, engine="openpyxl")
        print(f"[*] {cohort} predictions saved: {out_path}")
        if len(pred_df) >= 2:
            mvi_auc = roc_auc_score(pred_df["MVI_label"], pred_df["MVI_prob"])
            grade_auc = roc_auc_score(pred_df["PathologicalGrade_label"], pred_df["PathologicalGrade_prob"])
            print(f"    MVI AUC={mvi_auc:.4f}  Grade AUC={grade_auc:.4f}")


# ============================================================
# Entry point
# ============================================================
def parse_args():
    parser = argparse.ArgumentParser(
        description="Fine-tune custom pretrained DINOv2 on HCC CT volumes (MVI + Grade)")
    parser.add_argument("--cohort_a_excel", type=str, required=True,
                        help="Path to RJ clinical Excel file")
    parser.add_argument("--cohort_a_ap_dir", type=str, required=True,
                        help="Directory of RJ AP phase .nii.gz volumes")
    parser.add_argument("--cohort_a_pv_dir", type=str, required=True,
                        help="Directory of RJ PV phase .nii.gz volumes")
    parser.add_argument("--cohort_b_excel", type=str, required=True,
                        help="Path to XY clinical Excel file")
    parser.add_argument("--cohort_b_ap_dir", type=str, required=True,
                        help="Directory of XY AP phase .nii.gz volumes")
    parser.add_argument("--cohort_b_pv_dir", type=str, required=True,
                        help="Directory of XY PV phase .nii.gz volumes")
    parser.add_argument("--pretrained_weights", type=str, required=True,
                        help="Path to custom pretrained DINOv2 checkpoint (.pth)")
    parser.add_argument("--output_dir", type=str, default="./output_finetune_custom",
                        help="Output directory for checkpoints and predictions")
    parser.add_argument("--batch_size", type=int, default=4)
    parser.add_argument("--lr_list", type=str, default="3e-5,1e-4,3e-4",
                        help="Comma-separated list of learning rates to sweep")
    parser.add_argument("--weight_decay", type=float, default=1e-4)
    parser.add_argument("--epochs", type=int, default=10)
    parser.add_argument("--patience", type=int, default=3)
    return parser.parse_args()


def main():
    args = parse_args()
    cfg = Config(args)

    os.makedirs(cfg.OUTPUT_DIR, exist_ok=True)
    os.makedirs(cfg.TRAIN_LOG_DIR, exist_ok=True)

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"[*] Device: {device}")
    print(f"[*] LR sweep: {cfg.LR_LIST}")

    train_ds, val_ds, ext1_ds, ext2_ds = prepare_data_splits(cfg)

    loader_kwargs = dict(
        batch_size=cfg.BATCH_SIZE, collate_fn=my_collate,
        num_workers=0, pin_memory=True)

    dataloaders = {
        'train': data.DataLoader(train_ds, shuffle=True, **loader_kwargs),
        'val': data.DataLoader(val_ds, shuffle=False, **loader_kwargs),
    }
    # Held-out evaluation sets (not used during training)
    ext1_loader = data.DataLoader(ext1_ds, shuffle=False, **loader_kwargs)
    ext2_loader = data.DataLoader(ext2_ds, shuffle=False, **loader_kwargs)

    all_summary_rows = []

    for idx, lr in enumerate(cfg.LR_LIST, start=1):
        print("\n" + "#" * 70)
        print(f"[*] LR experiment {idx}/{len(cfg.LR_LIST)}: LR = {lr:.2e}")
        print("#" * 70)

        set_lr(cfg, lr)
        lr_tag, weight_dir, train_log_dir = prepare_run_dirs(cfg, lr)

        model = MultiTaskMILModel(cfg).to(device)
        summary_row = train_model(
            cfg=cfg, model=model, dataloaders=dataloaders,
            device=device, weight_dir=weight_dir,
            train_log_dir=train_log_dir, lr_tag=lr_tag)

        all_summary_rows.append(summary_row)
        update_lr_summary(train_log_dir, summary_row)

        del model
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    print("\nAll LR experiments completed.")
    print(f"[*] Output root: {cfg.OUTPUT_DIR}")
    print(f"[*] Logs: {cfg.TRAIN_LOG_DIR}")

    best_result = select_best_lr_and_save_report(cfg.TRAIN_LOG_DIR)

    if best_result is not None:
        best_lr = best_result["overall_best_lr"]["lr"]
        print(f"\n[*] Recommended best LR = {best_lr}")

        print("\n" + "=" * 70)
        print("  Running inference on best model")
        print("=" * 70)
        run_inference(cfg, best_result, device)


if __name__ == '__main__':
    main()
